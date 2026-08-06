#!/usr/bin/env python3
"""
Danh mục Module → Nhóm chức năng → Chức năng cho ERP bán theo module.
Nguồn: Digione, HORECA (24 phân hệ), Bahung HRM/LMS, chuẩn ERP thị trường.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cay_chuc_nang_data import TREE

OUT = Path(__file__).resolve().parent / "Danh_muc_Module_Chuc_nang_ERP_v3.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
MOD_FILLS = {
    "Nền tảng": PatternFill("solid", fgColor="D6EAF8"),
    "Nhân sự & Đào tạo": PatternFill("solid", fgColor="D5F5E3"),
    "Bán hàng & Khách hàng": PatternFill("solid", fgColor="FCF3CF"),
    "Chuỗi cung ứng": PatternFill("solid", fgColor="FADBD8"),
    "Sản xuất & Dịch vụ": PatternFill("solid", fgColor="E8DAEF"),
    "Tài chính": PatternFill("solid", fgColor="D4E6F1"),
    "Quản trị & Báo cáo": PatternFill("solid", fgColor="F5EEF8"),
}
MUST_FILL = PatternFill("solid", fgColor="F8CBAD")
SELL_FILL = PatternFill("solid", fgColor="C8E6C9")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

# Ma_Module, Ten_Module, Lop, Ban_rieng, Bat_buoc_kem, Mo_ta, Nguon_tham_chieu
MODULES = [
    ("SYS", "Hệ thống nền tảng", "Nền tảng", "Không (luôn kèm)", "—",
     "Auth, RBAC, tenant, cấu hình, menu, file, thông báo, nhắn tin realtime, audit, license.",
     "Digione System; HORECA DATA/ADM; chuẩn Identity & Access"),
    ("HRM", "Quản trị nhân sự", "Nhân sự & Đào tạo", "Có", "SYS",
     "Tổ chức, hồ sơ, tuyển dụng, ca/chấm công, tăng cường, nghỉ việc, lương.",
     "Digione HRM; Bahung HRM; chuẩn HCM"),
    ("LMS", "Đào tạo (LMS)", "Nhân sự & Đào tạo", "Có", "SYS",
     "Khóa học online/offline, mentor, thi, chứng chỉ, khảo sát.",
     "Digione LMS; Bahung LMS; HORECA OPS-03/04"),
    ("CRM", "CRM & Bán hàng", "Bán hàng & Khách hàng", "Có", "SYS",
     "Marketing, lead, cơ hội, sales online/offline, báo giá, đơn hàng, CSKH.",
     "Digione CRM; HORECA CRM-01..06"),
    ("POS", "POS bán lẻ", "Bán hàng & Khách hàng", "Có", "SYS",
     "Bán tại quầy, ca thu ngân, khuyến mại, thanh toán, đồng bộ tồn/BOM món.",
     "Digione POS; HORECA OPS-07/POS"),
    ("PUR", "Mua hàng", "Chuỗi cung ứng", "Có", "SYS",
     "NCC, yêu cầu mua, báo giá NCC, PO, nhận hàng, đối soát mua.",
     "HORECA ERP-01; chuẩn Procurement"),
    ("INV", "Kho & Tồn kho", "Chuỗi cung ứng", "Có", "SYS",
     "Đa kho, nhập/xuất/chuyển, lô/HSD/FEFO, giữ hàng, kiểm kê.",
     "HORECA ERP-02; chuẩn Inventory/WMS"),
    ("LOG", "Giao vận", "Chuỗi cung ứng", "Có", "SYS + INV",
     "Lệnh giao, COD, đối soát ship, theo dõi vận đơn.",
     "HORECA ERP-03 Ship"),
    ("MFG", "Sản xuất", "Sản xuất & Dịch vụ", "Có", "SYS + INV",
     "BOM, lệnh SX, định mức, tiêu hao NVL, thành phẩm, QC cơ bản.",
     "HORECA OPS-06 Rang; chuẩn Manufacturing"),
    ("FSM", "Dịch vụ kỹ thuật (Field Service)", "Sản xuất & Dịch vụ", "Có", "SYS + CRM",
     "Ticket bảo hành/sửa chữa, lịch KTV, linh kiện, SLA, nghiệm thu.",
     "HORECA OPS-01 FSM"),
    ("PJM", "Quản lý dự án / Setup", "Sản xuất & Dịch vụ", "Có", "SYS",
     "Dự án setup trọn gói, tiến độ, chi phí, nghiệm thu, gắn CRM/kho.",
     "HORECA OPS-02 Setup"),
    ("FIN", "Tài chính – Kế toán", "Tài chính", "Có", "SYS",
     "Sổ cái, công nợ phải thu/trả, thu chi, hóa đơn, kỳ kế toán, báo cáo TC.",
     "Digione FIN; HORECA ERP-04"),
    ("AST", "Quản lý tài sản", "Tài chính", "Có", "SYS",
     "Tài sản cố định, CCDC, khấu hao, bàn giao, kiểm kê TS.",
     "HORECA ERP-05; chuẩn Fixed Assets"),
    ("WF", "Công việc & Phê duyệt", "Quản trị & Báo cáo", "Có (hoặc kèm SYS)", "SYS",
     "Task, ticket nội bộ, workflow duyệt đa cấp, mẫu quy trình.",
     "HORECA ADM-02/03"),
    ("BI", "Báo cáo & BI", "Quản trị & Báo cáo", "Có", "SYS",
     "Dashboard KPI, báo cáo chuẩn theo module đã mua, cảnh báo.",
     "HORECA ADM-04; chuẩn BI"),
    ("PRT", "Cổng khách hàng / đối tác", "Bán hàng & Khách hàng", "Có", "SYS + CRM",
     "Portal KH xem đơn, công nợ, ticket; portal NCC/đối tác (tùy gói).",
     "Digione CPortal"),
]


def build_functions():
    """Trả về list (ma_mod, ten_mod, ma_nhom, ten_nhom, ma_cn, ten_cn, mo_ta, uu_tien)."""
    rows = []
    ten_by_ma = {m[0]: m[1] for m in MODULES}
    for ma_mod in (m[0] for m in MODULES):
        for nhom_code, ten_nhom, funcs in TREE[ma_mod]:
            for i, item in enumerate(funcs, 1):
                ten, mota, uu = item[0], item[1], item[2] if len(item) > 2 else "Cao"
                rows.append(
                    (
                        ma_mod,
                        ten_by_ma[ma_mod],
                        f"{ma_mod}-{nhom_code}",
                        ten_nhom,
                        f"{ma_mod}-{nhom_code}-{i:02d}",
                        ten,
                        mota,
                        uu,
                    )
                )
    return rows


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb = Workbook()

    # --- Sheet Hướng dẫn ---
    ws0 = wb.active
    ws0.title = "00_Huong_dan"
    ws0["A1"] = "DANH MỤC MODULE – NHÓM CHỨC NĂNG – CHỨC NĂNG (ERP BÁN THEO MODULE)"
    ws0["A1"].font = TITLE_FONT
    lines = [
        "",
        "Mục đích: Chốt phạm vi nghiệp vụ trước khi thiết kế source code.",
        "Cấu trúc: Module → Nhóm chức năng → Chức năng (mỗi chức năng = 1 dòng).",
        "",
        "Nguyên tắc bán module:",
        "1) SYS luôn đi kèm mọi gói (nền tảng license, user, phân quyền).",
        "2) Module nghiệp vụ bán riêng; cột 'Phụ thuộc' ghi module nên có kèm.",
        "3) Chức năng BI/Dashboard của từng module nằm trong module đó; BI là lớp tổng hợp.",
        "4) Trạng thái: Đề xuất — chờ bạn duyệt/chỉnh trước khi làm SRS chi tiết.",
        "",
        "Nguồn tham chiếu: Digione (System/HRM/LMS/CRM/POS/FIN/CPortal), HORECA 24 phân hệ,",
        "Bahung HRM-LMS (~334 UC), chuẩn ERP thị trường (Finance, HRM, CRM, Inventory, Procurement,",
        "Manufacturing, FSM, BI...).",
        "",
        "Cách duyệt: Sheet 01 xem danh mục module → Sheet 02 lọc theo Ma_Module → ghi chú cột Ghi_chu.",
        "Sau khi chốt: tách từng module thành tài liệu chi tiết trong thư mục 01. Modules.",
        "",
        "Phiên bản dữ liệu: v3 — chức năng GENERIC (chuẩn ERP bán được), không gắn 1 khách/ngành cụ thể.",
        "Cấu hình theo ngành (F&B, sản xuất, phân phối…) sẽ làm ở giai đoạn triển khai / template, không hard-code vào danh mục gốc.",
    ]
    for i, t in enumerate(lines, 2):
        ws0.cell(i, 1, t)
    ws0.column_dimensions["A"].width = 110

    # --- Sheet Modules ---
    ws1 = wb.create_sheet("01_Danh_muc_Module")
    headers1 = [
        "STT", "Ma_Module", "Ten_Module", "Lop", "Ban_rieng", "Phu_thuoc",
        "Mo_ta", "Nguon_tham_chieu", "Trang_thai", "Ghi_chu",
    ]
    for i, h in enumerate(headers1, 1):
        ws1.cell(1, i, h)
    style_header(ws1, 1, len(headers1))
    for idx, m in enumerate(MODULES, 1):
        vals = [idx, m[0], m[1], m[2], m[3], m[4], m[5], m[6], "Đề xuất", ""]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(idx + 1, c, v)
            cell.alignment = WRAP
            cell.border = THIN
            cell.font = Font(name="Calibri", size=10)
            if c == 4 and m[2] in MOD_FILLS:
                cell.fill = MOD_FILLS[m[2]]
            if c == 5 and str(v).startswith("Có"):
                cell.fill = SELL_FILL
            if m[0] == "SYS" and c == 2:
                cell.fill = MUST_FILL
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{1 + len(MODULES)}"
    autosize(ws1, [6, 12, 28, 22, 22, 16, 55, 40, 12, 25])

    # --- Sheet Functions ---
    ws2 = wb.create_sheet("02_Cay_chuc_nang")
    headers2 = [
        "STT", "Ma_Module", "Ten_Module", "Ma_Nhom", "Ten_Nhom",
        "Ma_CN", "Ten_Chuc_nang", "Mo_ta", "Uu_tien", "Trang_thai", "Ghi_chu",
    ]
    for i, h in enumerate(headers2, 1):
        ws2.cell(1, i, h)
    style_header(ws2, 1, len(headers2))
    funcs = build_functions()
    for idx, r in enumerate(funcs, 1):
        ma_mod, ten_mod, ma_nhom, ten_nhom, ma_cn, ten_cn, mota, uu = r
        vals = [idx, ma_mod, ten_mod, ma_nhom, ten_nhom, ma_cn, ten_cn, mota, uu, "Đề xuất", ""]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(idx + 1, c, v)
            cell.alignment = WRAP
            cell.border = THIN
            cell.font = Font(name="Calibri", size=10)
            if c == 9 and v == "Bắt buộc":
                cell.fill = MUST_FILL
        # color module col by layer
        layer = next(m[2] for m in MODULES if m[0] == ma_mod)
        if layer in MOD_FILLS:
            ws2.cell(idx + 1, 2).fill = MOD_FILLS[layer]
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{1 + len(funcs)}"
    autosize(ws2, [6, 12, 26, 12, 28, 16, 42, 48, 12, 12, 22])
    dv = DataValidation(type="list", formula1='"Bắt buộc,Cao,Trung bình,Thấp"', allow_blank=True)
    ws2.add_data_validation(dv)
    dv.add(f"I2:I{1 + len(funcs)}")
    dv2 = DataValidation(
        type="list",
        formula1='"Đề xuất,Giữ,Sửa,Gộp,Tách,Bỏ,Phase 1,Phase sau"',
        allow_blank=True,
    )
    ws2.add_data_validation(dv2)
    dv2.add(f"J2:J{1 + len(funcs)}")

    # --- Sheet Summary ---
    ws3 = wb.create_sheet("03_Tom_tat")
    ws3["A1"] = "TỔNG HỢP SỐ LƯỢNG"
    ws3["A1"].font = TITLE_FONT
    h = ["Ma_Module", "Ten_Module", "So_Nhom", "So_Chuc_nang", "Bat_buoc", "Cao", "Khac"]
    for i, x in enumerate(h, 1):
        ws3.cell(3, i, x)
    style_header(ws3, 3, len(h))
    from collections import defaultdict

    by = defaultdict(lambda: {"nhom": set(), "n": 0, "must": 0, "high": 0, "other": 0, "ten": ""})
    for r in funcs:
        ma, ten, ma_n, _, _, _, _, uu = r
        by[ma]["ten"] = ten
        by[ma]["nhom"].add(ma_n)
        by[ma]["n"] += 1
        if uu == "Bắt buộc":
            by[ma]["must"] += 1
        elif uu == "Cao":
            by[ma]["high"] += 1
        else:
            by[ma]["other"] += 1
    row = 4
    total_n = total_g = 0
    for m in MODULES:
        ma = m[0]
        d = by[ma]
        vals = [ma, d["ten"], len(d["nhom"]), d["n"], d["must"], d["high"], d["other"]]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row, c, v)
            cell.border = THIN
            cell.alignment = WRAP
        total_n += d["n"]
        total_g += len(d["nhom"])
        row += 1
    ws3.cell(row + 1, 1, "TỔNG")
    ws3.cell(row + 1, 1).font = Font(bold=True)
    ws3.cell(row + 1, 3, total_g)
    ws3.cell(row + 1, 4, total_n)
    ws3.cell(row + 3, 1, f"Tổng module: {len(MODULES)} | Tổng nhóm: {total_g} | Tổng chức năng: {total_n}")
    ws3.cell(row + 3, 1).font = Font(bold=True, color="1F4E79")
    autosize(ws3, [12, 30, 12, 14, 12, 10, 10])

    # --- Sheet Gói bán ---
    ws4 = wb.create_sheet("04_Goi_ban_de_xuat")
    ws4["A1"] = "GỢI Ý GÓI BÁN (có thể chỉnh sau khi chốt nghiệp vụ)"
    ws4["A1"].font = TITLE_FONT
    packs = [
        ("Gói", "Module gồm", "Đối tượng khách"),
        ("Starter", "SYS + CRM + FIN", "Công ty thương mại nhỏ, cần bán & sổ sách"),
        ("Retail", "SYS + POS + INV + FIN (+ CRM)", "Chuỗi cửa hàng / F&B bán lẻ"),
        ("Distribution", "SYS + CRM + PUR + INV + LOG + FIN", "Phân phối, giao hàng, công nợ"),
        ("Service", "SYS + CRM + FSM + INV + FIN (+ PRT)", "Bảo hành / kỹ thuật hiện trường"),
        ("People", "SYS + HRM + LMS (+ FIN lương)", "Nhà máy / chuỗi cần nhân sự & đào tạo"),
        ("Manufacture", "SYS + INV + MFG + PUR + FIN (+ HRM)", "Sản xuất / rang xay / gia công"),
        ("Full ERP", "Tất cả module", "Doanh nghiệp đa mô hình (kiểu HORECA)"),
    ]
    for r, pack in enumerate(packs, 3):
        for c, v in enumerate(pack, 1):
            cell = ws4.cell(r, c, v)
            cell.border = THIN
            cell.alignment = WRAP
            if r == 3:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
    autosize(ws4, [16, 55, 50])

    # --- Sheet Nguồn ---
    ws5 = wb.create_sheet("05_Nguon_tham_chieu")
    ws5["A1"] = "ÁNH XẠ NGUỒN → MODULE SẢN PHẨM"
    ws5["A1"].font = TITLE_FONT
    map_rows = [
        ("Nguồn", "Phân hệ / nội dung gốc", "Module sản phẩm"),
        ("Digione", "System-Solution", "SYS"),
        ("Digione", "HRM-Solution", "HRM"),
        ("Digione", "LMS-Solution", "LMS"),
        ("Digione", "CRM-Solution", "CRM"),
        ("Digione", "POS-Solution", "POS"),
        ("Digione", "FIN-Solution", "FIN"),
        ("Digione", "CPortal-Solution", "PRT"),
        ("HORECA", "DATA-01 Dữ liệu nền tảng", "SYS (+ master trong từng module)"),
        ("HORECA", "CRM-01..06 Marketing→Sales Admin", "CRM"),
        ("HORECA", "ERP-01 Mua hàng", "PUR"),
        ("HORECA", "ERP-02 Kho", "INV"),
        ("HORECA", "ERP-03 Ship/Giao vận", "LOG"),
        ("HORECA", "ERP-04 Kế toán–Tài chính", "FIN"),
        ("HORECA", "ERP-05 Tài sản", "AST"),
        ("HORECA", "ERP-06 Hợp đồng", "CRM / PJM / FIN (tùy loại HĐ)"),
        ("HORECA", "OPS-01 Kỹ thuật FSM", "FSM"),
        ("HORECA", "OPS-02 Dự án Setup", "PJM"),
        ("HORECA", "OPS-03/04 Đào tạo Offline + LMS", "LMS"),
        ("HORECA", "OPS-05 R&D", "PJM / MFG (phase sau chi tiết)"),
        ("HORECA", "OPS-06 Sản xuất rang", "MFG"),
        ("HORECA", "OPS-07 Chuỗi 80plus + POS", "POS (+ INV/HRM vận hành CH)"),
        ("HORECA", "ADM-01 Nhân sự–Payroll", "HRM"),
        ("HORECA", "ADM-02/03 Task & Engine duyệt", "WF"),
        ("HORECA", "ADM-04 BI Dashboard", "BI"),
        ("Bahung", "14 service HRM-LMS / 334 UC", "HRM + LMS (chi tiết hóa sau)"),
        ("Thị trường ERP", "Finance, Procurement, Inventory, Manufacturing, CRM, HRM, BI…", "Ánh xạ các module trên"),
    ]
    for r, rowv in enumerate(map_rows, 3):
        for c, v in enumerate(rowv, 1):
            cell = ws5.cell(r, c, v)
            cell.border = THIN
            cell.alignment = WRAP
            if r == 3:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
    autosize(ws5, [16, 55, 40])

    wb.save(OUT)
    print(f"Wrote: {OUT}")
    print(f"Modules: {len(MODULES)} | Groups: {total_g} | Functions: {total_n}")


if __name__ == "__main__":
    main()
